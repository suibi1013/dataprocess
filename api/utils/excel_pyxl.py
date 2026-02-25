from typing import List, Dict, Any, Union, Tuple, Optional
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

def _col_to_index_0based(col: Union[str, int],max_col: int) -> int:
    """将列索引转换为 0-based 的列索引
    :param col: 列索引（1-based 或字母表示）
    :param max_col: 最大列索引（用于验证,1-based）
    :return: 0-based 的列索引
    """    
    if isinstance(col, str):
        if col.strip()=="": col=max_col
        else:
            # 先尝试将字符串转换为整数
            if not col.isalpha(): 
                col = int(col)
    if isinstance(col, int):
        # 验证列索引是否在有效范围内,写入时，允许超出最大索引
        # if col > max_col: raise ValueError(f"Column index must be <= {max_col}")
        if col==0: raise ValueError(f"Column index can't be 0")
        # 转换为 0-based 索引
        if col < 0:col=max_col+col
        else:col-=1
        return col
    elif isinstance(col, str):        
        col = col.upper().strip()
        idx = 0
        for ch in col:
            if not ch.isalpha(): raise ValueError(f"Invalid column letter: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx-1   
# 索引颜色列表（openpyxl 内置 COLOR_INDEX 前16项）
COLOR_INDEX = [
    "00000000", "00FFFFFF", "00FF0000", "0000FF00", "000000FF", "00FFFF00",
    "00FF00FF", "0000FFFF", "00800000", "00FF8080", "0080FF80", "008080FF",
    "00FFFF80", "00FF80FF", "0080FFFF", "00C0C0C0"
]
DEFAULT_THEME_COLORS = {
    0: "000000",  # Dark1
    1: "FFFFFF",  # Light1
    2: "4472C4",  # Accent1
    3: "ED7D31",  # Accent2
    4: "A5A5A5",  # Accent3
    5: "FFC000",  # Accent4
    6: "5B9BD5",  # Accent5
    7: "70AD47",  # Accent6
    8: "0000FF",  # Hyperlink
    9: "800080",  # Followed Hyperlink
}
# 获取主题RGB颜色的函数
def _get_theme_rgb(theme, tint,workbook):
    """从主题索引获取RGB颜色"""
    #"""将 Excel 主题色 + 色调转换为 ARGB 字符串（如 'FF00FF00'）"""
    try:
        base_rgb = DEFAULT_THEME_COLORS[theme]
        # 验证 base_rgb 是 6 位十六进制
        if not base_rgb or not re.match(r'^[0-9A-Fa-f]{6}$', base_rgb):
            return None

        # 应用 tint（色调调整）
        def apply_tint(rgb_hex, tint_value):
            r = int(rgb_hex[0:2], 16)
            g = int(rgb_hex[2:4], 16)
            b = int(rgb_hex[4:6], 16)

            if tint_value < 0:
                # Darken
                r = int(r * (1 + tint_value))
                g = int(g * (1 + tint_value))
                b = int(b * (1 + tint_value))
            else:
                # Lighten
                r = int(r * (1 - tint_value) + 255 * tint_value)
                g = int(g * (1 - tint_value) + 255 * tint_value)
                b = int(b * (1 - tint_value) + 255 * tint_value)

            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            return f"{r:02X}{g:02X}{b:02X}"

        rgb = apply_tint(base_rgb, tint)
        return f"{rgb}".upper()  # Alpha=FF (不透明)
    except Exception:
        return None

def _get_color_rgb(cell_color, workbook=None): 
    """ 
    Safely extract RGB hex string (without alpha) from openpyxl color object. 
    Returns None if no color. 
    """ 
    if not cell_color:
        return None 
    
    color_rgb=None
    if (cell_color.indexed == 0 or cell_color.indexed is None) and cell_color.rgb is None:
        return None
     
    if cell_color.type == 'rgb' and cell_color.rgb:
        # 提取RGB值，去掉alpha通道
        color_rgb = cell_color.rgb[2:] if len(cell_color.rgb) == 8 else cell_color.rgb
    elif cell_color.type == 'theme' and workbook is not None:
        color_rgb= _get_theme_rgb(cell_color.theme, getattr(cell_color, 'tint', 0.0),workbook) 
    elif cell_color.type == 'indexed':
        idx = cell_color.indexed 
        if idx < len(COLOR_INDEX):
            indexed_rgb = COLOR_INDEX[idx] 
            if indexed_rgb and len(indexed_rgb) == 8:
                color_rgb = indexed_rgb[2:]
    return color_rgb

# ==================== 样式转换函数 ====================
def _convert_color_to_dict(color, workbook=None) -> Dict[str, Any]:
    """将Color对象转换为可序列化的字典，只保留关键颜色信息"""
    if not color:
        return None
    
    color_dict = {'rgb':None,'theme':None,'tint':None,'type':None}
    # 使用安全的方式获取RGB值
    rgb_value = _get_color_rgb(color,workbook)
    if rgb_value:
        color_dict['rgb'] = rgb_value
    
    # 仍然保留其他颜色属性
    try:
        if hasattr(color, 'theme'):
            # 只在theme是有效整数时添加
            theme_val = color.theme
            if isinstance(theme_val, int):
                color_dict['theme'] = theme_val
    except Exception:
        pass
    
    try:
        if hasattr(color, 'tint'):
            color_dict['tint'] = color.tint
    except Exception:
        pass
    
    try:
        if hasattr(color, 'type'):
            color_dict['type'] = color.type
    except Exception:
        pass
    return color_dict

def _convert_font_to_dict(font, workbook=None) -> Dict[str, Any]:
    """将Font对象转换为可序列化的字典"""
    return {
        'name': font.name,
        'size': font.size,
        'bold': font.bold,
        'italic': font.italic,
        'underline': font.underline,
        'strike': font.strike,
        'color': _convert_color_to_dict(font.color,workbook),
        'vertAlign': font.vertAlign,
        'scheme': font.scheme
    }

def _convert_fill_to_dict(fill,workbook=None) -> Dict[str, Any]:
    """将Fill对象转换为可序列化的字典"""
    if hasattr(fill, 'patternType'):
        return {
            'patternType': fill.patternType,
            'fgColor': _convert_color_to_dict(fill.fgColor,workbook),
            'bgColor': _convert_color_to_dict(fill.bgColor,workbook)
        }
    return {}

def _convert_side_to_dict(side,workbook=None) -> Dict[str, Any]:
    """将Side对象转换为可序列化的字典"""
    if not side:
        return {'style': None, 'color': None}
    return {
        'style': side.style,
        'color': _convert_color_to_dict(side.color,workbook)
    }

def _convert_border_to_dict(border,workbook=None) -> Dict[str, Any]:
    """将Border对象转换为可序列化的字典"""
    return {
        'left': _convert_side_to_dict(border.left,workbook),
        'right': _convert_side_to_dict(border.right,workbook),
        'top': _convert_side_to_dict(border.top,workbook),
        'bottom': _convert_side_to_dict(border.bottom,workbook),
        'diagonal': _convert_side_to_dict(border.diagonal,workbook),
        'diagonal_direction': border.diagonal_direction,
        'outline': border.outline,
        'vertical': border.vertical,
        'horizontal': border.horizontal
    }

def _convert_alignment_to_dict(alignment,workbook=None) -> Dict[str, Any]:
    """将Alignment对象转换为可序列化的字典"""
    return {
        'horizontal': alignment.horizontal,
        'vertical': alignment.vertical,
        'text_rotation': alignment.text_rotation,
        'wrap_text': alignment.wrap_text,
        'shrink_to_fit': alignment.shrink_to_fit,
        'indent': alignment.indent,
        'relativeIndent': alignment.relativeIndent,
        'justifyLastLine': alignment.justifyLastLine,
        'readingOrder': alignment.readingOrder
    }

def _convert_protection_to_dict(protection) -> Dict[str, Any]:
    """将Protection对象转换为可序列化的字典"""
    return {
        'locked': protection.locked,
        'hidden': protection.hidden
    }
# ==================== 步骤一：读取区域 → 转为 A1 起点的局部矩阵 ====================
def read_excel_range_data(
    file_path: str,
    sheet_name: str,
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
    start_col: Optional[Union[str, int]] = None,
    end_col: Optional[Union[str, int]] = None,
    include_styles: bool = False,
) -> Dict[str, Any]:
    """
    读取 Excel 指定区域的数据（含样式和值）。
    
    如果未指定行列范围，则自动读取整个工作表的有效区域。
    
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param start_row: 起始行（1-based），默认自动
    :param end_row: 结束行（1-based），默认自动
    :param start_col: 起始列（如 "A" 或 1），默认自动
    :param end_col: 结束列（如 "Z" 或 26），默认自动
    :param include_styles: 是否包含样式信息（如字体、填充、边框等），默认 False
    :return: 字典，包含 'data'（二维列表，每个单元格为 {'value', 'style_id' 或 'style', 'abs_pos'}）和 'style_map'（样式ID到样式字典的映射）
    """
    # 先加载 value-only 版本以获取最大行列（更快）
    workbook = load_workbook(file_path, data_only=True)
    ws_value = workbook[sheet_name]
    
    # 转换列为 1-based 索引，处理负数索引（-1 表示末尾）    
    n_rows = ws_value.max_row
    n_cols = ws_value.max_column
    start_row = _col_to_index_0based(start_row,n_rows)+1
    start_col_idx = _col_to_index_0based(start_col,n_cols)+1
    end_row = _col_to_index_0based(end_row,n_rows)+1
    end_col_idx = _col_to_index_0based(end_col,n_cols)+1
    
    # 校验范围
    if start_row > end_row or start_col_idx > end_col_idx:
        workbook.close()
        raise ValueError(f"Start must be <= End.start_row({start_row}), end_row({end_row}), start_col({start_col}), end_col({end_col})")
    
    # 获取样式
    wb_style=None
    ws_style=None
    if include_styles:
        # 加载带样式的 workbook
        wb_style = load_workbook(file_path, data_only=False)
        ws_style = wb_style[sheet_name]
    
    result = []
    style_map = {}
    style_id_counter = 1
    style_cache = {}
    
    for r in range(start_row, end_row + 1):
        row = []
        for c in range(start_col_idx, end_col_idx + 1):
            value_cell: Cell = ws_value.cell(row=r, column=c)
            cell_data = {
                'value': value_cell.value,
                'abs_pos': [r,c],
            }
            #读取单元格样式
            if include_styles:
                style_cell: Cell = ws_style.cell(row=r, column=c)
                # 生成完整样式字典
                style_dict = {
                    'font': _convert_font_to_dict(style_cell.font,workbook),
                    'fill': _convert_fill_to_dict(style_cell.fill,workbook) if style_cell.fill.fill_type else {},
                    'border': _convert_border_to_dict(style_cell.border,workbook),
                    'alignment': _convert_alignment_to_dict(style_cell.alignment,workbook),
                    'number_format': style_cell.number_format,
                    'protection': _convert_protection_to_dict(style_cell.protection),
                }
                
                # 使用样式字典的字符串表示作为缓存键
                style_key = str(style_dict)
                if style_key not in style_cache:
                    # 新样式，分配ID并添加到映射表
                    style_id = style_id_counter
                    style_id_counter += 1
                    style_cache[style_key] = style_id
                    style_map[style_id] = style_dict
                else:
                    # 已存在的样式，使用缓存的ID
                    style_id = style_cache[style_key]
                # 在单元格数据中存储样式ID
                cell_data['style_id'] = style_id
            else:
                # 不包含样式
                cell_data['style_id'] = None
            row.append(cell_data)
        result.append(row)
    
    if include_styles:
        wb_style.close()
    workbook.close()
    
    # 返回包含数据和样式映射的字典
    return {
        'data': result,
        'style_map': style_map
    }
from openpyxl.utils import range_boundaries
def read_excel_range_data_by_range(
    file_path: str,
    sheet_name: str,
    excel_cell_range: Optional[str] = None,
    include_styles: bool = False,
) -> Dict[str, Any]:
    """
    读取 Excel 某个区域数据。
    
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param excel_cell_range: 区域范围，如 "A1:D10"
    :param include_styles: 是否包含样式信息（如字体、填充、边框等），默认 False
    :return: 字典，包含 'data'（二维列表，每个单元格为 {'value', 'style_id' 或 'style', 'abs_pos'}）和 'style_map'（样式ID到样式字典的映射）
    """
    min_col, min_row, max_col, max_row = range_boundaries(excel_cell_range)
    return read_excel_range_data(
        file_path=file_path,
        sheet_name=sheet_name,
        start_row=min_row,
        end_row=max_row,
        start_col=min_col,
        end_col=max_col,
        include_styles=include_styles,
    )