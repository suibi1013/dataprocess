from typing import List, Dict, Any, Union, Tuple, Optional
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

def _col_to_index_0based(col: Union[str, int],max_col: int) -> int:
    """将列索引转换为 0-based 的列索引
    :param col: 列索引（1-based 或字母表示）
    :param max_col: 最大列索引（用于验证,1-based）
    :return: 0-based 的列索引
    """    
    if isinstance(col, str):
        if col.strip()=="": col=max_col
        else:
            # 先尝试将字符串转换为整数
            if not col.isalpha(): 
                col = int(col)
    if isinstance(col, int):
        # 验证列索引是否在有效范围内,写入时，允许超出最大索引
        # if col > max_col: raise ValueError(f"Column index must be <= {max_col}")
        if col==0: raise ValueError(f"Column index can't be 0")
        # 转换为 0-based 索引
        if col < 0:col=max_col+col
        else:col-=1
        return col
    elif isinstance(col, str):        
        col = col.upper().strip()
        idx = 0
        for ch in col:
            if not ch.isalpha(): raise ValueError(f"Invalid column letter: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx-1   
# 索引颜色列表（openpyxl 内置 COLOR_INDEX 前16项）
COLOR_INDEX = [
    "00000000", "00FFFFFF", "00FF0000", "0000FF00", "000000FF", "00FFFF00",
    "00FF00FF", "0000FFFF", "00800000", "00FF8080", "0080FF80", "008080FF",
    "00FFFF80", "00FF80FF", "0080FFFF", "00C0C0C0"
]

# 获取主题RGB颜色的函数
def get_theme_rgb(theme, tint,workbook):
    """从主题索引获取RGB颜色"""
    #"""将 Excel 主题色 + 色调转换为 ARGB 字符串（如 'FF00FF00'）"""
    try:
        clr_scheme = workbook.theme.themeElements.clrScheme
        # 映射 theme 索引到 clrScheme 属性
        theme_colors = [
            clr_scheme.lt1.val,   # 0: Light 1 (usually white)
            clr_scheme.dk1.val,   # 1: Dark 1 (usually black)
            clr_scheme.lt2.val,   # 2: Light 2
            clr_scheme.dk2.val,   # 3: Dark 2
            clr_scheme.accent1.val,
            clr_scheme.accent2.val,
            clr_scheme.accent3.val,
            clr_scheme.accent4.val,
            clr_scheme.accent5.val,
            clr_scheme.accent6.val,
        ]
        if theme < len(theme_colors):
            base_rgb = theme_colors[theme]
        else:
            return None

        # 验证 base_rgb 是 6 位十六进制
        if not base_rgb or not re.match(r'^[0-9A-Fa-f]{6}$', base_rgb):
            return None

        # 应用 tint（色调调整）
        def apply_tint(rgb_hex, tint_value):
            r = int(rgb_hex[0:2], 16)
            g = int(rgb_hex[2:4], 16)
            b = int(rgb_hex[4:6], 16)

            if tint_value < 0:
                # Darken
                r = int(r * (1 + tint_value))
                g = int(g * (1 + tint_value))
                b = int(b * (1 + tint_value))
            else:
                # Lighten
                r = int(r * (1 - tint_value) + 255 * tint_value)
                g = int(g * (1 - tint_value) + 255 * tint_value)
                b = int(b * (1 - tint_value) + 255 * tint_value)

            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            return f"{r:02X}{g:02X}{b:02X}"

        rgb = apply_tint(base_rgb, tint)
        return f"FF{rgb}".upper()  # Alpha=FF (不透明)
    except Exception:
        return None

def get_color_rgb(cell_color, workbook=None): 
    """ 
    Safely extract RGB hex string (without alpha) from openpyxl color object. 
    Returns None if no color. 
    """ 
    if not cell_color:
        return None 
    
    color_rgb=None
    if (cell_color.indexed == 0 or cell_color.indexed is None) and cell_color.rgb is None:
        return None
     
    if cell_color.type == 'rgb' and cell_color.rgb:
        # 提取RGB值，去掉alpha通道
        color_rgb = cell_color.rgb[2:] if len(cell_color.rgb) == 8 else cell_color.rgb
    elif cell_color.type == 'theme' and workbook is not None:
        color_rgb= get_theme_rgb(workbook, cell_color.theme, getattr(cell_color, 'tint', 0.0)) 
    elif cell_color.type == 'indexed':
        idx = cell_color.indexed 
        if idx < len(COLOR_INDEX):
            indexed_rgb = COLOR_INDEX[idx] 
            if indexed_rgb and len(indexed_rgb) == 8:
                color_rgb = indexed_rgb[2:]
    return color_rgb

# ==================== 样式转换函数 ====================
def _convert_color_to_dict(color, workbook=None) -> Dict[str, Any]:
    """将Color对象转换为可序列化的字典，只保留关键颜色信息"""
    if not color:
        return None
    
    color_dict = {}
    # 使用安全的方式获取RGB值
    rgb_value = get_color_rgb(color,workbook)
    if rgb_value:
        color_dict['rgb'] = rgb_value
    
    # 仍然保留其他颜色属性
    try:
        if hasattr(color, 'theme'):
            # 只在theme是有效整数时添加
            theme_val = color.theme
            if isinstance(theme_val, int):
                color_dict['theme'] = theme_val
    except Exception:
        pass
    
    try:
        if hasattr(color, 'tint'):
            color_dict['tint'] = color.tint
    except Exception:
        pass
    
    try:
        if hasattr(color, 'type'):
            color_dict['type'] = color.type
    except Exception:
        pass
    # 如果没有获取到任何有效属性，返回None
    return color_dict if color_dict else None

def _convert_font_to_dict(font, workbook=None) -> Dict[str, Any]:
    """将Font对象转换为可序列化的字典"""
    return {
        'name': font.name,
        'size': font.size,
        'bold': font.bold,
        'italic': font.italic,
        'underline': font.underline,
        'strike': font.strike,
        'color': _convert_color_to_dict(font.color,workbook),
        'vertAlign': font.vertAlign,
        'scheme': font.scheme
    }

def _convert_fill_to_dict(fill,workbook=None) -> Dict[str, Any]:
    """将Fill对象转换为可序列化的字典"""
    if hasattr(fill, 'patternType'):
        return {
            'patternType': fill.patternType,
            'fgColor': _convert_color_to_dict(fill.fgColor,workbook),
            'bgColor': _convert_color_to_dict(fill.bgColor,workbook)
        }
    return {}

def _convert_side_to_dict(side,workbook=None) -> Dict[str, Any]:
    """将Side对象转换为可序列化的字典"""
    if not side:
        return {'style': None, 'color': None}
    return {
        'style': side.style,
        'color': _convert_color_to_dict(side.color,workbook)
    }

def _convert_border_to_dict(border,workbook=None) -> Dict[str, Any]:
    """将Border对象转换为可序列化的字典"""
    return {
        'left': _convert_side_to_dict(border.left,workbook),
        'right': _convert_side_to_dict(border.right,workbook),
        'top': _convert_side_to_dict(border.top,workbook),
        'bottom': _convert_side_to_dict(border.bottom,workbook),
        'diagonal': _convert_side_to_dict(border.diagonal,workbook),
        'diagonal_direction': border.diagonal_direction,
        'outline': border.outline,
        'vertical': border.vertical,
        'horizontal': border.horizontal
    }

def _convert_alignment_to_dict(alignment,workbook=None) -> Dict[str, Any]:
    """将Alignment对象转换为可序列化的字典"""
    return {
        'horizontal': alignment.horizontal,
        'vertical': alignment.vertical,
        'text_rotation': alignment.text_rotation,
        'wrap_text': alignment.wrap_text,
        'shrink_to_fit': alignment.shrink_to_fit,
        'indent': alignment.indent,
        'relativeIndent': alignment.relativeIndent,
        'justifyLastLine': alignment.justifyLastLine,
        'readingOrder': alignment.readingOrder
    }

def _convert_protection_to_dict(protection) -> Dict[str, Any]:
    """将Protection对象转换为可序列化的字典"""
    return {
        'locked': protection.locked,
        'hidden': protection.hidden
    }
# ==================== 步骤一：读取区域 → 转为 A1 起点的局部矩阵 ====================
def read_excel_range_data(
    file_path: str,
    sheet_name: str,
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
    start_col: Optional[Union[str, int]] = None,
    end_col: Optional[Union[str, int]] = None,
    include_styles: bool = False,
) -> Dict[str, Any]:
    """
    读取 Excel 指定区域的数据（含样式和值）。
    
    如果未指定行列范围，则自动读取整个工作表的有效区域。
    
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param start_row: 起始行（1-based），默认自动
    :param end_row: 结束行（1-based），默认自动
    :param start_col: 起始列（如 "A" 或 1），默认自动
    :param end_col: 结束列（如 "Z" 或 26），默认自动
    :param include_styles: 是否包含样式信息（如字体、填充、边框等），默认 False
    :return: 字典，包含 'data'（二维列表，每个单元格为 {'value', 'style_id' 或 'style', 'abs_pos'}）和 'style_map'（样式ID到样式字典的映射）
    """
    # 先加载 value-only 版本以获取最大行列（更快）
    workbook = load_workbook(file_path, data_only=True)
    ws_value = workbook[sheet_name]
    
    # 转换列为 1-based 索引，处理负数索引（-1 表示末尾）    
    n_rows = ws_value.max_row
    n_cols = ws_value.max_column
    start_row = _col_to_index_0based(start_row,n_rows)+1
    start_col_idx = _col_to_index_0based(start_col,n_cols)+1
    end_row = _col_to_index_0based(end_row,n_rows)+1
    end_col_idx = _col_to_index_0based(end_col,n_cols)+1
    
    # 校验范围
    if start_row > end_row or start_col_idx > end_col_idx:
        workbook.close()
        raise ValueError(f"Start must be <= End.start_row({start_row}), end_row({end_row}), start_col({start_col}), end_col({end_col})")
    
    # 获取样式
    wb_style=None
    ws_style=None
    if include_styles:
        # 加载带样式的 workbook
        wb_style = load_workbook(file_path, data_only=False)
        ws_style = wb_style[sheet_name]
    
    result = []
    style_map = {}
    style_id_counter = 1
    style_cache = {}
    
    for r in range(start_row, end_row + 1):
        row = []
        for c in range(start_col_idx, end_col_idx + 1):
            value_cell: Cell = ws_value.cell(row=r, column=c)
            cell_data = {
                'value': value_cell.value,
                'abs_pos': [r,c],
            }
            #读取单元格样式
            if include_styles:
                style_cell: Cell = ws_style.cell(row=r, column=c)
                # 生成完整样式字典
                style_dict = {
                    'font': _convert_font_to_dict(style_cell.font,workbook),
                    'fill': _convert_fill_to_dict(style_cell.fill,workbook) if style_cell.fill.fill_type else {},
                    'border': _convert_border_to_dict(style_cell.border,workbook),
                    'alignment': _convert_alignment_to_dict(style_cell.alignment,workbook),
                    'number_format': style_cell.number_format,
                    'protection': _convert_protection_to_dict(style_cell.protection),
                }
                
                # 使用样式字典的字符串表示作为缓存键
                style_key = str(style_dict)
                if style_key not in style_cache:
                    # 新样式，分配ID并添加到映射表
                    style_id = style_id_counter
                    style_id_counter += 1
                    style_cache[style_key] = style_id
                    style_map[style_id] = style_dict
                else:
                    # 已存在的样式，使用缓存的ID
                    style_id = style_cache[style_key]
                # 在单元格数据中存储样式ID
                cell_data['style_id'] = style_id
            else:
                # 不包含样式
                cell_data['style_id'] = None
            row.append(cell_data)
        result.append(row)
    
    if include_styles:
        wb_style.close()
    workbook.close()
    
    # 返回包含数据和样式映射的字典
    return {
        'data': result,
        'style_map': style_map
    }

import xlwings as xw
from typing import List, Dict, Any, Union, Tuple
import os
import time

def _col_to_index_0based(col: Union[str, int],max_col: int) -> int:
    """将列索引转换为 0-based 的列索引
    :param col: 列索引（1-based 或字母表示）
    :param max_col: 最大列索引（用于验证,1-based）
    :return: 0-based 的列索引
    """
    if isinstance(col, str):
        if col.strip()=="": col=max_col
        else:
            # 先尝试将字符串转换为整数
            if not col.isalpha(): 
                col = int(col)
    if isinstance(col, int):
        # 验证列索引是否在有效范围内,写入时，允许超出最大索引
        # if col > max_col: raise ValueError(f"Column index must be <= {max_col}")
        if col==0: raise ValueError(f"Column index can't be 0")
        # 转换为 0-based 索引
        if col < 0:col=max_col+col
        else:col-=1
        return col
    elif isinstance(col, str):        
        col = col.upper().strip()
        idx = 0
        for ch in col:
            if not ch.isalpha(): raise ValueError(f"Invalid column letter: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx-1   
def _find_contiguous_regions(cells: List[Tuple[int, int]]) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    """查找单元格列表中的连续矩形区域
    
    优化：支持多行连续区域的合并，减少区域数量
    
    :param cells: 单元格列表，格式为 [(row, col), ...]（0-based）
    :return: 连续区域列表，每个区域为 ((start_row, start_col), (end_row, end_col))
    """
    if not cells:
        return []
    
    # 按行排序，然后按列排序
    sorted_cells = sorted(cells, key=lambda x: (x[0], x[1]))
    
    # 首先按行合并连续列区域
    row_regions = []
    current_row = None
    current_start_col = None
    current_end_col = None
    
    for row, col in sorted_cells:
        if current_row is None:
            # 开始新行
            current_row = row
            current_start_col = col
            current_end_col = col
        elif row == current_row:
            # 同一行
            if col == current_end_col + 1:
                # 连续列
                current_end_col = col
            else:
                # 不连续，结束当前区域并开始新区域
                row_regions.append((current_row, current_start_col, current_end_col))
                current_start_col = col
                current_end_col = col
        else:
            # 新行，结束当前区域
            row_regions.append((current_row, current_start_col, current_end_col))
            current_row = row
            current_start_col = col
            current_end_col = col
    
    # 添加最后一个区域
    if current_row is not None:
        row_regions.append((current_row, current_start_col, current_end_col))
    
    # 现在合并多行连续的相同列范围区域
    if not row_regions:
        return []
    
    merged_regions = []
    # 按行和列范围排序
    sorted_row_regions = sorted(row_regions, key=lambda x: (x[1], x[2], x[0]))
    
    current_start_row = None
    current_end_row = None
    current_start_col = None
    current_end_col = None
    
    for row, start_col, end_col in sorted_row_regions:
        if current_start_row is None:
            # 开始新区域
            current_start_row = row
            current_end_row = row
            current_start_col = start_col
            current_end_col = end_col
        else:
            # 检查是否可以合并到当前区域
            # 条件：相同的列范围，并且行是连续的
            if (start_col == current_start_col and end_col == current_end_col and 
                row == current_end_row + 1):
                # 合并到当前区域
                current_end_row = row
            else:
                # 无法合并，结束当前区域并开始新区域
                merged_regions.append(((current_start_row, current_start_col), (current_end_row, current_end_col)))
                current_start_row = row
                current_end_row = row
                current_start_col = start_col
                current_end_col = end_col
    
    # 添加最后一个区域
    if current_start_row is not None:
        merged_regions.append(((current_start_row, current_start_col), (current_end_row, current_end_col)))
    
    return merged_regions


def _apply_style_to_cells(ws, style: Dict[str, Any], regions: List[Tuple[Tuple[int, int], Tuple[int, int]]]):
    """批量应用样式到一组连续区域（0-based 行列坐标）
    
    :param ws: 工作表对象
    :param style: 样式字典
    :param regions: 连续区域列表，每个区域为 ((start_row, start_col), (end_row, end_col))
    """
    if not regions:
        return
    
    # 对每个连续区域应用样式
    for (start_row, start_col), (end_row, end_col) in regions:
        # 创建连续区域的 Range 对象（1-based）
        rng = ws.range((start_row + 1, start_col + 1), (end_row + 1, end_col + 1))

        # === 字体 ===
        font_style = style.get('font', {})
        if font_style:
            if 'name' in font_style:
                rng.font.name = font_style['name']
            if 'size' in font_style:
                rng.font.size = font_style['size']
            if 'bold' in font_style:
                rng.font.bold = bool(font_style['bold'])
            if 'italic' in font_style:
                rng.font.italic = bool(font_style.get('italic', False))
            if 'underline' in font_style:
                rng.font.underline = font_style.get('underline') is not None
            if 'strike' in font_style:
                rng.font.strikethrough = bool(font_style.get('strike', False))
            color = font_style.get('color')
            if color and color.get('type') == 'rgb':
                rgb_str = color.get('rgb', '000000').lstrip('#').ljust(6, '0')[:6]
                try:
                    r = int(rgb_str[0:2], 16)
                    g = int(rgb_str[2:4], 16)
                    b = int(rgb_str[4:6], 16)
                    rng.font.color = (r, g, b)
                except Exception:
                    pass  # 忽略无效颜色

        # === 背景色 ===
        fill = style.get('fill', {})
        fg_color = fill.get('fgColor')
        if fg_color and fg_color.get('type') == 'rgb':
            rgb_str = fg_color.get('rgb', 'FFFFFF').lstrip('#').ljust(6, '0')[:6]
            try:
                r = int(rgb_str[0:2], 16)
                g = int(rgb_str[2:4], 16)
                b = int(rgb_str[4:6], 16)
                rng.color = (r, g, b)
            except Exception:
                pass

        # === 对齐 ===
        alignment = style.get('alignment', {})
        if alignment:
            vert_map = {'center': -4108, 'top': -4160, 'bottom': -4107}
            horz_map = {'center': -4108, 'left': -4131, 'right': -4152}

            api_rng = ws.api.Range(rng.address)
            if 'vertical' in alignment:
                v_align = alignment['vertical']
                api_rng.VerticalAlignment = vert_map.get(v_align, -4108)
            if 'horizontal' in alignment:
                h_align = alignment['horizontal']
                api_rng.HorizontalAlignment = horz_map.get(h_align, -4131)
            if 'wrap_text' in alignment:
                api_rng.WrapText = bool(alignment['wrap_text'])

        # === 边框（详细处理）===
        border = style.get('border', {})
        if border:
            api_rng = ws.api.Range(rng.address)
            
            # 定义边框边缘对应的常量
            edge_map = {
                'left': 7,      # xlEdgeLeft
                'right': 10,    # xlEdgeRight
                'top': 8,       # xlEdgeTop
                'bottom': 9     # xlEdgeBottom
            }
            
            # 处理每条边框
            for side_name, edge in edge_map.items():
                side_data = border.get(side_name, {})
                border_style = side_data.get('style')
                color = side_data.get('color')
                
                # 只有当样式或颜色存在时才设置边框
                if border_style or color:
                    try:
                        # 获取当前边缘的Border对象，减少重复调用
                        border_edge = api_rng.Borders(edge)
                        # 一次性设置边框属性
                        if border_style:
                            # 映射样式名称到Excel常量
                            line_style_map = {
                                'none':         -4142,   # xlLineStyleNone
                                'continuous':   1,       # 实线（默认）
                                'thin':         1,       # 通常等同于 continuous
                                'medium':       1,
                                'thick':        1,
                                'dashed':       2,       # 虚线
                                'dotted':       3,       # 点线
                                'dash_dot':     4,       # 点划线（短划+点）
                                'dash_dot_dot': 5,       # 双点划线（短划+两点）
                                'slant_dash_dot': 13,    # 斜点划线（Excel 特有）
                                'double':       6,       # 双线
                            }
                            border_edge.LineStyle = line_style_map.get(border_style, 1)
                            
                            if border_style == 'medium':
                                border_edge.Weight = 3  # xlMedium
                            elif border_style == 'thick':
                                border_edge.Weight = 4  # xlThick
                            else:
                                border_edge.Weight = 2  # xlThin
                        else:
                            border_edge.LineStyle = 1  # 默认实线 (xlContinuous)
                            border_edge.Weight = 2  # 默认细线
                        
                        # 设置边框颜色
                        if color and color.get('type') == 'rgb':
                            rgb_str = color.get('rgb', '000000').lstrip('#').ljust(6, '0')[:6]
                            try:
                                # Excel VBA使用BGR格式的长整型颜色值
                                r = int(rgb_str[0:2], 16)
                                g = int(rgb_str[2:4], 16)
                                b = int(rgb_str[4:6], 16)
                                bgr_color = b + (g << 8) + (r << 16)
                                border_edge.Color = bgr_color
                            except Exception as e:
                                print(f"警告: 设置边框颜色时出错: {e}")
                    except Exception as e:
                        print(f"警告: 设置边框时出错: {e}")


def write_data_to_existing_excel(
    data: Dict[str, Any],
    file_path: str,
    sheet_name: str,
    start_row: int = 1,
    start_col: Union[str, int] = 1,
    write_style: bool = True,
    chunk_size: int = 10000
) -> None:
    """
    高效分块写入带样式的结构化数据到现有 Excel 文件（同步分块处理）

    :param data: 表格数据，包含 'data'（二维列表，每个元素为 {"value":"xx","abs_pos":[],"style_id":0}）和 'style_map'（样式ID到样式字典的映射）
    :param file_path: 目标 Excel 文件路径（必须存在）
    :param sheet_name: 工作表名
    :param start_row: 起始行（1-based，默认第1行）
    :param start_col: 起始列（1-based 整数 或 字母，如 1 或 'A'）
    :param write_style: 是否写入样式
    :param chunk_size: 分块大小，默认10000行
    """
    # 分块处理逻辑
    if not data or not data.get('data') or not data['data'][0]:
        return

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    total_start_time = time.time()
    # 获取数据和样式映射
    excel_data = data['data']
    style_map = data.get('style_map', {})
    total_rows = len(excel_data)
    n_cols = len(excel_data[0])

    # 转换起始位置为 0-based
    start_row_0 = _col_to_index_0based(start_row, total_rows)
    start_col_0 = _col_to_index_0based(start_col, n_cols)

    # 数据分块
    def chunk_data_local(data, chunk_size):
        if not data:
            return []
        chunks = []
        total_rows = len(data)
        for i in range(0, total_rows, chunk_size):
            end = min(i + chunk_size, total_rows)
            chunk = {
                'data': data[i:end],
                'start_row': i,
                'end_row': end - 1
            }
            chunks.append(chunk)
        return chunks
    
    chunks = chunk_data_local(excel_data, chunk_size)
    total_chunks = len(chunks)
    print(f"数据分块完成，共 {total_chunks} 个块，每个块最多 {chunk_size} 行")

    # 分块处理
    app = None
    wb = None
    ws = None
    processed_rows = 0
    style_groups_applied = 0

    try:
        # 初始化 Excel 应用
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        # 打开工作簿
        wb = app.books.open(file_path)

        # 确保工作表存在
        if sheet_name not in wb.sheet_names:
            ws = wb.sheets.add(name=sheet_name)
        else:
            ws = wb.sheets[sheet_name]

        # 处理每个数据块
        for chunk_idx, chunk in enumerate(chunks):
            chunk_data = chunk['data']
            chunk_start = chunk['start_row']
            chunk_end = chunk['end_row']
            chunk_rows = chunk_end - chunk_start + 1
            
            print(f"处理第 {chunk_idx + 1}/{total_chunks} 块，行范围: {chunk_start + 1}-{chunk_end + 1}，共 {chunk_rows} 行")

            # 构建当前块的值矩阵和样式映射
            value_matrix = []
            # 使用样式ID作为键的样式映射
            cell_style_map: Dict[int, List[Tuple[int, int]]] = {}

            for i, row in enumerate(chunk_data):
                value_row = []
                for j, cell in enumerate(row):
                    if cell is None:
                        value_row.append("")
                        continue

                    value = cell.get('value', "")
                    value_row.append(value)

                    # 使用 style_id 从 style_map 获取样式
                    if write_style:
                        style_id = cell.get('style_id')
                        if style_id is not None and style_id in style_map:
                            target_row = start_row_0 + chunk_start + i
                            target_col = start_col_0 + j
                            if style_id not in cell_style_map:
                                cell_style_map[style_id] = []
                            cell_style_map[style_id].append((target_row, target_col))
                value_matrix.append(value_row)

            # 写入当前块的值
            if value_matrix:
                current_block_start_row = start_row_0 + chunk_start + 1  # 转换为 1-based
                current_block_start_col = start_col_0 + 1  # 转换为 1-based
                ws.range((current_block_start_row, current_block_start_col)).value = value_matrix
                print(f"  写入值完成")

            # 应用当前块的样式
            if write_style and cell_style_map:
                # 预处理样式映射，查找连续矩形区域
                style_regions_map: Dict[int, List[Tuple[Tuple[int, int], Tuple[int, int]]]] = {}
                for style_id, cells in cell_style_map.items():
                    regions = _find_contiguous_regions(cells)
                    style_regions_map[style_id] = regions

                # 应用样式
                for style_id, regions in style_regions_map.items():
                    try:
                        style = style_map.get(style_id, {})
                        if style:
                            _apply_style_to_cells(ws, style, regions)
                    except Exception as e:
                        print(f"  应用样式时出错: {e}")
                
                style_groups_applied += len(cell_style_map)
                print(f"  应用样式完成，共 {len(cell_style_map)} 个样式组")

            # 每个块写完后保存
            wb.save()
            print(f"  保存完成")
            processed_rows += chunk_rows

    except Exception as e:
        raise RuntimeError(f"写入 Excel 失败: {e}") from e
    finally:
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except Exception:
            pass  # 忽略清理错误

    total_time = time.time() - total_start_time
    print(f"写入完成，共处理 {total_rows} 行数据，耗时 {round(total_time, 2)} 秒")
    return

from typing import List, Dict, Any
def structure_custom_table_with_style_dynamic(
    data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    结构化表格，支持：
      - 动态元信息列（WPO_SEG4 = ...）
      - GroupName 行识别（仅第一列有值）
      - 空行为分割标识，处理多个数据区域
    
    :param data:表格数据，包含 'data'（二维列表，每个元素为 {"value":"xx","abs_pos":[],"style_id":0}）和 'style_map'（样式ID到样式字典的映射）
    :return: 结构化表格，包含 'data'（二维列表，每个元素为 {"value":"xx","abs_pos":[],"style_id":0}）和 'style_map'（样式ID到样式字典的映射）
    """
    if not data or not data.get('data') or not data['data'][0]:
        return data
    
    # 获取数据
    excel_data = data['data']
    
    # 辅助函数：判断是否为全空行
    def is_empty_row(row):
        if not row:
            return True
        for cell in row:
            val = cell['value']
            if val is not None and str(val).strip() != "":
                return False
        return True
    
    # Step 1: 分割数据为多个区域（以空行为分隔）
    data_regions = []
    current_region = []
    
    for row in excel_data:
        if is_empty_row(row):
            # 遇到空行，结束当前区域并开始新区域
            if current_region:
                data_regions.append(current_region)
                current_region = []
        else:
            current_region.append(row)
    
    # 添加最后一个区域
    if current_region:
        data_regions.append(current_region)
    
    if not data_regions:
        return data
    
    # Step 2: 处理每个数据区域
    all_result_rows = []
    header_cells = None
    
    for region in data_regions:
        if not region or not region[0]:
            continue
        
        # 找标题行（第一个第二列非空）
        header_row_index = None
        for i, row in enumerate(region):
            if len(row) >= 2 and row[1]['value'] is not None and str(row[1]['value']).strip() != "":
                header_row_index = i
                break
        if header_row_index is None:
            continue  # 跳过无标题行的区域
        
        original_header_row = region[header_row_index]
        
        # 解析元信息行
        dynamic_meta = []
        for row in region[:header_row_index]:
            if not row or row[0]['value'] is None:
                continue
            s = str(row[0]['value']).strip()
            if "=" in s:
                key, val_part = s.split("=", 1)
                key = key.strip()
                value = val_part.split("\\")[0].strip()
                dynamic_meta.append((key, {
                    'value': value,
                    'style': row[0].get('style', {}),
                    'abs_pos': row[0].get('abs_pos', [0, 0])
                }))
        
        dynamic_keys = [k for k, _ in dynamic_meta]
        measure_columns = original_header_row[1:]
        
        # 构建表头（仅第一次构建）
        if header_cells is None:
            header_cells = []
            for key in dynamic_keys:
                header_cells.append({'value': key, 'style': {}, 'abs_pos': [0, len(header_cells)]})
            for col in ["GroupName", "RowName"]:
                header_cells.append({'value': col, 'style': {}, 'abs_pos': [0, len(header_cells)]})
            header_cells.extend(measure_columns)
        
        # 处理数据行
        current_group_name_cell = {'value': '', 'style': {}, 'abs_pos': [0, 0]}
        
        for row in region[header_row_index + 1:]:
            if not row:
                continue
            
            # 判断是否为 GroupName 行：仅第一列有值，其余度量列为空
            is_group_row = True
            if len(row) <= 1:
                is_group_row = False
            else:
                for j in range(1, len(row)):
                    val = row[j]['value']
                    if val is not None and str(val).strip() != "":
                        is_group_row = False
                        break
            
            if is_group_row and len(row) > 0 and row[0]['value'] is not None and str(row[0]['value']).strip() != "":
                # 更新 GroupName
                current_group_name_cell = row[0]
            else:
                # 输出数据行
                new_row = []
                # 动态列
                for _, cell in dynamic_meta:
                    new_row.append(cell)
                # GroupName
                new_row.append(current_group_name_cell)
                # RowName
                row_name_cell = row[0] if len(row) > 0 else {'value': '', 'style': {}, 'abs_pos': [0, 0]}
                new_row.append(row_name_cell)
                # 度量列
                for j in range(1, len(original_header_row)):
                    if j < len(row):
                        new_row.append(row[j])
                    else:
                        new_row.append({
                            'value': None,
                            'style': {},
                            'abs_pos': [row_name_cell.get('abs_pos', [0, 0])[0], j]
                        })
                all_result_rows.append(new_row)
    
    # 构建最终结果
    if header_cells and all_result_rows:
        data['data'] = [header_cells] + all_result_rows
    
    return data



# 读取测试数据
file_path=r"E:\temp\aaa.xlsx"
sheet_name="Sheet1"
start_row="1"
end_row=""
start_col="A"
end_col=""
include_styles=False
import time
time_begin=time.time()
print(f"开始读取数据")
data=read_excel_range_data(file_path,sheet_name,start_row,end_row,start_col,end_col,include_styles)
import json
print(json.dumps(data, ensure_ascii=False))
# time_end=time.time()
# print(f"读取数据耗时: {time_end-time_begin} 秒")
# print(f"数据规模: {len(data['data'])} 行, {len(data['data'][0])} 列")
# import json
# print(json.dumps(data['style_map'], ensure_ascii=False))
# print(json.dumps(data['data'][:10], ensure_ascii=False))
# print(f"数据规模: {len(data['data'])} 行, {len(data['data'][0])} 列")
# print(f"数据规模:size: {len(data['style_map'])}")
# print(f"数据规模:data: {len(json.dumps(data['data'][:10], ensure_ascii=False))}")

# ptint(structure_custom_table_with_style_dynamic(data))

# import time
# import json
# def read_json_file(filepath: str) -> dict:
#     """
#     读取json文件信息，并转换为json对象        
#     Args:
#         filepath: json文件路径            
#     Returns:
#         dict: json对象
#     """
#     if not os.path.exists(filepath):
#         raise FileNotFoundError(f"文件不存在: {filepath}")
#     try:
#         with open(filepath, 'r', encoding='utf-8') as f:
#             return json.load(f)
#         # with open(filepath, 'rb') as f:
#         #     return orjson.loads(f.read())
#     except json.JSONDecodeError as e:
#         raise ValueError(f"JSON 格式无效 ({filepath}): {e}")
        
# file_path="E:\\projectcode\\dataprocess\\api\\config_infos/data_processes\\process_flows\\58bcae6f-02a5-4e7c-a934-b61d82d3689e\\debug\\node_1770278866993_dwcwa9wld.json"
# data=read_json_file(file_path)
# file_path=r"E:\temp\result.xlsx"
# time_begin=time.time()
# write_data_to_existing_excel(data,file_path,sheet_name,start_row,start_col,include_styles)
# time_end=time.time()
# print(f"写入数据耗时: {time_end-time_begin} 秒")

    


