from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from typing import List, Dict, Any, Union, Optional

def _col_to_index_0based(col: Union[str, int],max_col: int) -> int:
    """将列索引转换为 0-based 的列索引
    :param col: 列索引（1-based 或字母表示）
    :param max_col: 最大列索引（用于验证,1-based）
    :return: 0-based 的列索引
    """
    if isinstance(col, str):
        # 先尝试将字符串转换为整数
        if not col.isalpha(): 
            col = int(col)
    if isinstance(col, int):
        # 验证列索引是否在有效范围内
        if col > max_col: raise ValueError(f"Column index must be <= {max_col}")
        if col==0: raise ValueError(f"Column index can't be 0")
        # 转换为 0-based 索引
        if col < 0:col=max_col+col
        else:col-=1
        return col
    elif isinstance(col, str):        
        col = col.upper().strip()
        idx = 0
        for ch in col:
            if not ch.isalpha(): raise ValueError(f"Invalid column letter: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx-1      
# 索引颜色列表
COLOR_INDEX = [
    "00000000",  # 0: 黑色
    "00FFFFFF",  # 1: 白色
    "00FF0000",  # 2: 红色
    "0000FF00",  # 3: 绿色
    "000000FF",  # 4: 蓝色
    "00FFFF00",  # 5: 黄色
    "00FF00FF",  # 6: 洋红色
    "0000FFFF",  # 7: 青色
    "00800000",  # 8: 深灰色
    "00FF8080",  # 9: 浅红色
    "0080FF80",  # 10: 浅绿色
    "008080FF",  # 11: 浅蓝色
    "00FFFF80",  # 12: 浅黄色
    "00FF80FF",  # 13: 浅洋红色
    "0080FFFF",  # 14: 浅青色
    "00C0C0C0",  # 15: 浅灰色
    # 更多索引颜色...
]

# 获取主题RGB颜色的函数
def get_theme_rgb(theme, tint,workbook):
    """从主题索引获取RGB颜色"""
    #"""将 Excel 主题色 + 色调转换为 ARGB 字符串（如 'FF00FF00'）"""
    try:
        clr_scheme = workbook.theme.themeElements.clrScheme
        # 映射 theme 索引到 clrScheme 属性
        theme_colors = [
            clr_scheme.lt1.val,   # 0: Light 1 (usually white)
            clr_scheme.dk1.val,   # 1: Dark 1 (usually black)
            clr_scheme.lt2.val,   # 2: Light 2
            clr_scheme.dk2.val,   # 3: Dark 2
            clr_scheme.accent1.val,
            clr_scheme.accent2.val,
            clr_scheme.accent3.val,
            clr_scheme.accent4.val,
            clr_scheme.accent5.val,
            clr_scheme.accent6.val,
        ]
        if theme < len(theme_colors):
            base_rgb = theme_colors[theme]
        else:
            return None

        # 验证 base_rgb 是 6 位十六进制
        if not base_rgb or not re.match(r'^[0-9A-Fa-f]{6}$', base_rgb):
            return None

        # 应用 tint（色调调整）
        def apply_tint(rgb_hex, tint_value):
            r = int(rgb_hex[0:2], 16)
            g = int(rgb_hex[2:4], 16)
            b = int(rgb_hex[4:6], 16)

            if tint_value < 0:
                # Darken
                r = int(r * (1 + tint_value))
                g = int(g * (1 + tint_value))
                b = int(b * (1 + tint_value))
            else:
                # Lighten
                r = int(r * (1 - tint_value) + 255 * tint_value)
                g = int(g * (1 - tint_value) + 255 * tint_value)
                b = int(b * (1 - tint_value) + 255 * tint_value)

            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            return f"{r:02X}{g:02X}{b:02X}"

        rgb = apply_tint(base_rgb, tint)
        return f"FF{rgb}".upper()  # Alpha=FF (不透明)
    except Exception:
        return None

def get_color_rgb(cell_color, workbook=None): 
    """ 
    Safely extract RGB hex string (without alpha) from openpyxl color object. 
    Returns None if no color. 
    """ 
    if not cell_color:
        return None 
 
    if cell_color.type == 'rgb' and cell_color.rgb:
        # Remove alpha channel (first 2 chars) 
        return cell_color.rgb[2:] if len(cell_color.rgb) == 8 else cell_color.rgb 
    elif cell_color.type == 'theme' and workbook is not None:
        return get_theme_rgb(workbook, cell_color.theme, getattr(cell_color, 'tint', 0.0)) 
    elif cell_color.type == 'indexed':
        idx = cell_color.indexed 
        if idx < len(COLOR_INDEX):
            indexed_rgb = COLOR_INDEX[idx] 
            if indexed_rgb and len(indexed_rgb) == 8:
                return indexed_rgb[2:] 
    return None

# ==================== 样式转换函数 ====================
def _convert_color_to_dict(color, workbook=None) -> Dict[str, Any]:
    """将Color对象转换为可序列化的字典，只保留关键颜色信息"""
    if not color:
        return None
    
    color_dict = {}
    
    # 使用安全的方式获取RGB值
    rgb_value = get_color_rgb(color,workbook)
    if rgb_value:
        color_dict['rgb'] = rgb_value
    
    # 仍然保留其他颜色属性
    try:
        if hasattr(color, 'theme'):
            # 只在theme是有效整数时添加
            theme_val = color.theme
            if isinstance(theme_val, int):
                color_dict['theme'] = theme_val
    except Exception:
        pass
    
    try:
        if hasattr(color, 'tint'):
            color_dict['tint'] = color.tint
    except Exception:
        pass
    
    try:
        if hasattr(color, 'type'):
            color_dict['type'] = color.type
    except Exception:
        pass
    
    # 如果没有获取到任何有效属性，返回None
    return color_dict if color_dict else None

def _convert_font_to_dict(font, workbook=None) -> Dict[str, Any]:
    """将Font对象转换为可序列化的字典"""
    return {
        'name': font.name,
        'size': font.size,
        'bold': font.bold,
        'italic': font.italic,
        'underline': font.underline,
        'strike': font.strike,
        'color': _convert_color_to_dict(font.color,workbook),
        'vertAlign': font.vertAlign,
        'scheme': font.scheme
    }

def _convert_fill_to_dict(fill,workbook=None) -> Dict[str, Any]:
    """将Fill对象转换为可序列化的字典"""
    if hasattr(fill, 'patternType'):
        return {
            'patternType': fill.patternType,
            'fgColor': _convert_color_to_dict(fill.fgColor,workbook),
            'bgColor': _convert_color_to_dict(fill.bgColor,workbook)
        }
    return {}

def _convert_side_to_dict(side,workbook=None) -> Dict[str, Any]:
    """将Side对象转换为可序列化的字典"""
    if not side:
        return {'style': None, 'color': None}
    return {
        'style': side.style,
        'color': _convert_color_to_dict(side.color,workbook)
    }

def _convert_border_to_dict(border,workbook=None) -> Dict[str, Any]:
    """将Border对象转换为可序列化的字典"""
    return {
        'left': _convert_side_to_dict(border.left,workbook),
        'right': _convert_side_to_dict(border.right,workbook),
        'top': _convert_side_to_dict(border.top,workbook),
        'bottom': _convert_side_to_dict(border.bottom,workbook),
        'diagonal': _convert_side_to_dict(border.diagonal,workbook),
        'diagonal_direction': border.diagonal_direction,
        'outline': border.outline,
        'vertical': border.vertical,
        'horizontal': border.horizontal
    }

def _convert_alignment_to_dict(alignment,workbook=None) -> Dict[str, Any]:
    """将Alignment对象转换为可序列化的字典"""
    return {
        'horizontal': alignment.horizontal,
        'vertical': alignment.vertical,
        'text_rotation': alignment.text_rotation,
        'wrap_text': alignment.wrap_text,
        'shrink_to_fit': alignment.shrink_to_fit,
        'indent': alignment.indent,
        'relativeIndent': alignment.relativeIndent,
        'justifyLastLine': alignment.justifyLastLine,
        'readingOrder': alignment.readingOrder
    }

def _convert_protection_to_dict(protection) -> Dict[str, Any]:
    """将Protection对象转换为可序列化的字典"""
    return {
        'locked': protection.locked,
        'hidden': protection.hidden
    }

# ==================== 步骤一：读取区域 → 转为 A1 起点的局部矩阵 ====================
def read_excel_range_data(
    file_path: str,
    sheet_name: str,
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
    start_col: Optional[Union[str, int]] = None,
    end_col: Optional[Union[str, int]] = None,
    include_styles: bool = False,
) -> List[List[Dict[str, Any]]]:
    """
    读取 Excel 指定区域的数据（含样式和值）。
    
    如果未指定行列范围，则自动读取整个工作表的有效区域。
    
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param start_row: 起始行（1-based），默认自动
    :param end_row: 结束行（1-based），默认自动
    :param start_col: 起始列（如 "A" 或 1），默认自动
    :param end_col: 结束列（如 "Z" 或 26），默认自动
    :param include_styles: 是否包含样式信息（如字体、填充、边框等），默认 False
    :return: 二维列表，每个单元格为 {'value', 'style', 'abs_pos'}
    """
    # 先加载 value-only 版本以获取最大行列（更快）
    workbook = load_workbook(file_path, data_only=True)
    ws_value = workbook[sheet_name]
    
    # 转换列为 1-based 索引，处理负数索引（-1 表示末尾）    
    n_rows = ws_value.max_row
    n_cols = ws_value.max_column
    start_row = _col_to_index_0based(start_row,n_rows)+1
    start_col_idx = _col_to_index_0based(start_col,n_cols)+1
    end_row = _col_to_index_0based(end_row,n_rows)+1
    end_col_idx = _col_to_index_0based(end_col,n_cols)+1

    # 校验范围
    if start_row > end_row or start_col_idx > end_col_idx:
        workbook.close()
        raise ValueError(f"Start must be <= End.start_row({start_row}), end_row({end_row}), start_col({start_col}), end_col({end_col})")

    # 获取样式
    wb_style=None
    ws_style=None
    if include_styles:
        # 加载带样式的 workbook
        wb_style = load_workbook(file_path, data_only=False)
        ws_style = wb_style[sheet_name]

    result = []
    for r in range(start_row, end_row + 1):
        row = []
        for c in range(start_col_idx, end_col_idx + 1):
            value_cell: Cell = ws_value.cell(row=r, column=c)
            cell_data = {
                'value': value_cell.value,
                'style': {},
                'abs_pos': [r,c],
            }
            #读取单元格样式
            if include_styles:
                style_cell: Cell = ws_style.cell(row=r, column=c)
                cell_data['style']={
                    'font': _convert_font_to_dict(style_cell.font,workbook),
                    'fill': _convert_fill_to_dict(style_cell.fill,workbook),
                    'border': _convert_border_to_dict(style_cell.border,workbook),
                    'alignment': _convert_alignment_to_dict(style_cell.alignment,workbook),
                    'number_format': style_cell.number_format,
                    'protection': _convert_protection_to_dict(style_cell.protection),
                }
            row.append(cell_data)
        result.append(row)

    if include_styles:
        wb_style.close()
    workbook.close()

    return result

import pandas as pd
import re
from typing import List, Tuple, Dict, Any, Union

def _col_to_index_0based(col: Union[str, int],max_col: int) -> int:
    """将列索引转换为 0-based 的列索引
    :param col: 列索引（1-based 或字母表示）
    :param max_col: 最大列索引（用于验证,1-based）
    :return: 0-based 的列索引
    """
    if isinstance(col, str):
        # 先尝试将字符串转换为整数
        if not col.isalpha(): 
            col = int(col)
    if isinstance(col, int):
        # 验证列索引是否在有效范围内
        if col > max_col: raise ValueError(f"Column index must be <= {max_col}")
        if col==0: raise ValueError(f"Column index can't be 0")
        # 转换为 0-based 索引
        if col < 0:col=max_col+col
        else:col-=1
        return col
    elif isinstance(col, str):        
        col = col.upper().strip()
        idx = 0
        for ch in col:
            if not ch.isalpha(): raise ValueError(f"Invalid column letter: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx-1    
import re
import json
def make_regex(s: str) -> str:
    if not s:
        raise ValueError("字符串不能为空")
    # 转义特殊字符
    escaped_s = re.escape(s)
    if len(set(s))==1:
        # 纯字符,只需匹配后边不以第一位字符开头即可
        escaped_s0 = re.escape(s[0])
        return f"^{escaped_s}(?!{escaped_s0})"
    return f"^{escaped_s}(?!{escaped_s})"
def group_rows_by_feature_segments(
    data: List[List[Dict[str, Any]]],
    split_column: Union[str, int],
    beginstr: Union[str, List[str]]=None,
    pattern_regex: Union[str, List[str]]=None
) -> List[Tuple[int, int]]:
    """
    根据指定列的值是否匹配任一正则表达式，将数据行分段。

    参数:
        df: 输入表格（不含 header）
        split_column: 分段列索引(1-based,如 "A" 或 1)
        beginstr: 开头字符串特征列表,格式为["xx1","xx2"]，如["   ","  "]，表示以两个或三个空格开头的特征值
        pattern_regex: 正则表达式特征列表,如["[\\u4e00-\\u9fff]"],表示包含中文的特征值

    返回:
        List[Tuple[int, int]]: 分组范围 [(start_row, end_row), ...]（Excel 行号）
    """
    if not data or not split_column :
        return []
    print('beginstr',beginstr)
    print('pattern_regex',pattern_regex)
    # 转换字符串为数组类型
    # 只在beginstr是字符串时才进行JSON解析
    if beginstr and isinstance(beginstr, str):
        try:
            beginstr=json.loads(beginstr)
        except Exception as e:
            raise ValueError(f"beginstr数据格式错误: {e}")
    
    # 只在pattern_regex是字符串时才进行JSON解析
    if pattern_regex and isinstance(pattern_regex, str):
        try:
            pattern_regex=json.loads(pattern_regex)
        except Exception as e:
            raise ValueError(f"pattern_regex数据格式错误: {e}")

    # 提取列名和数据值
    headers = data[0]
    # 从每个列头字典中提取value作为实际列名
    column_names = [header['value'] for header in headers]
    # 将数据行转换为仅包含值的列表
    rows_values = [[cell['value'] for cell in row] for row in data[1:]]
    df = pd.DataFrame(rows_values, columns=column_names)

    # 获取分段列索引（0-based）
    num_rows = len(data)
    num_cols = len(data[0]) if num_rows > 0 else 0    
    split_column_index = _col_to_index_0based(split_column, num_cols)
    # 获取列名
    split_column_name = column_names[split_column_index]

    # 构建正则表达式列表
    patterns=[]
    if beginstr:
        # 统一为列表
        beginstr=[beginstr] if isinstance(beginstr, str) else beginstr
        patterns.extend([make_regex(s) for s in beginstr])
    if pattern_regex:
        # 统一为列表
        pattern_regex=[pattern_regex] if isinstance(pattern_regex, str) else pattern_regex
        patterns.extend(pattern_regex)
    print('patterns',pattern_regex,patterns)
    if not patterns:
        return []
    # 预编译所有正则
    compiled_patterns = [re.compile(p) for p in patterns]

    segment_starts: List[int] = []

    # 遍历每一行
    for idx, value in enumerate(df[split_column_name]):
        if not isinstance(value, str):
            continue

        # 只要匹配任意一个正则，就记录为起始行
        for cp in compiled_patterns:
            if cp.search(value):                
                segment_starts.append(idx+1)
                break  # 匹配一个即可，避免重复添加

    if not segment_starts:
        return []

    segments: List[Tuple[int, int]] = []
    n_rows = len(df)

    for i, start_idx in enumerate[any](segment_starts):
        if i == len(segment_starts) - 1:
            end_idx = n_rows - 1
        else:
            end_idx = segment_starts[i + 1] - 1

        excel_start = start_idx + 1
        excel_end = end_idx + 1
        segments.append((excel_start, excel_end))

    return segments

from typing import List, Dict, Any, Union, Callable, Optional

def data_grouprow_split_by_column(
    data: List[List[Dict[str, Any]]],    
    split_column: Union[str, int],    
    beginstr: Union[str, List[str]]=None,
    pattern_regex: Union[str, List[str]]=None,
    new_col_header_name:str="new_col",
    new_col_position="left",
    charecter_cell_replace_value:str=None
) -> List[List[Dict[str, Any]]]:
    """
    数据行分组按列拆分，返回新的二维数据。

    :param data: 二维数据
    :param split_column: 特征值列索引(1-based,如 "A" 或 1)
    :param segments: 分组范围 [(start_row, end_row), ...]（Excel 行号）
    :param beginstr: 开头字符串特征列表,格式为["xx1","xx2"]，如["   ","  "]，表示以两个或三个空格开头的特征值
    :param pattern_regex: 正则表达式特征列表,如["[\\u4e00-\\u9fff]"],表示包含中文的特征值
    :param new_col_header_name: 新列标题名称
    :param new_col_position: 新列位置,相对于拆分列的位置，左侧或右侧（left/right）, [{"value": "left", "label":"左侧"},{"value": "right", "label":"右侧"}]
    :param charecter_cell_replace_value: 分组内特征单元格值替换为（可选）
    :return: 新的二维数据
    """
    if not data:
        # 空数据，返回空
        return []

    num_rows = len(data)
    num_cols = len(data[0]) if num_rows > 0 else 0

    # 查找新列插入位置
    insert_at_col_index = _col_to_index_0based(split_column,num_cols)+1
    if new_col_position=="right":
        insert_at_col_index = insert_at_col_index
    else:
        insert_at_col_index = insert_at_col_index-1
    
    # 将数据行，按特征分段
    segments = group_rows_by_feature_segments(
        data=data,
        split_column=split_column,
        beginstr=beginstr,  # 如两个空格开头，第三个不是空格
        pattern_regex=pattern_regex
    )
    print("分组范围（Excel 行号）:",segments)
    # 构建新数据
    new_data = []
    # 添加标题行
    new_row = []
    new_row.extend(data[0][:insert_at_col_index])
    # 新列标题也应该是一个字典对象
    new_col_header = {
        'value': new_col_header_name,
        'style': {},
        'abs_pos': [1, insert_at_col_index + 1]  # 1-based 逻辑位置
    }
    new_row.append(new_col_header)
    new_row.extend(data[0][insert_at_col_index:])
    new_data.append(new_row)

    # 处理数据行
    for start_row, end_row in segments:
        segment_rows = data[start_row-1:end_row]
        # 获取特征值        
        charecter_cell = segment_rows[0][insert_at_col_index]
        # 提取特征值的value
        charecter_cell_value = charecter_cell['value']
        if charecter_cell_replace_value:
            segment_rows[0][insert_at_col_index]['value'] = charecter_cell_replace_value
        # 处理数据行
        for row_idx, row in enumerate(segment_rows):  
            new_row = []
            new_row.extend(row[:insert_at_col_index])            

            # 新单元格（abs_pos 使用逻辑位置：行号不变，列设为 insert_at_col_index+1）
            # 单元格样式暂未设置
            new_cell = {
                'value': charecter_cell_value,
                'style': {},
                'abs_pos': [row_idx + 2, insert_at_col_index + 1]  # 1-based 逻辑位置
            }
            new_row.append(new_cell)
            new_row.extend(row[insert_at_col_index:])
            new_data.append(new_row)

    return new_data

file_path="E:/temp/欣和酱油测试.xlsx"
sheet_name="Sheet2"
data = read_excel_range_data(file_path=file_path, sheet_name=sheet_name,start_row=1,end_row=469,start_col="A",end_col="F", include_styles=False)

# for index, row in enumerate(data):
#     print(row)
#     if index>10:
#         break

# Step 2: 按特征分段
# beginstrarray = ["   ","  "]
# beginstrarray ="[\"   \",\"  \"]"
beginstrarray =""
pattern_regex = ["[\\u4e00-\\u9fff]"]

# Step 3:数据行分组按列拆分
newdata=data_grouprow_split_by_column(
    data=data,
    split_column="D",
    beginstr=beginstrarray,
    pattern_regex=pattern_regex,
    new_col_header_name="品牌",
    new_col_position="left",
    charecter_cell_replace_value="酱油总体"
)
print("新数据行分组按列拆分后:")
for index, row in enumerate(newdata):
    print(index,row)
    if index>10:
        break